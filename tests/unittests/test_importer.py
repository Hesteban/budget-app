"""
Unit tests for budget/importer.py

Pure logic tests — no Supabase connection needed.
Synthetic .xlsx bytes come from the account_xlsx_bytes fixture (conftest.py).
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
import pandas as pd
import pytest

from budget.importer import (
    _clean_amount,
    _parse_date,
    detect_format,
    df_to_records,
    parse_bank_file,
    parse_bank_file_bulk,
    parse_account_format,
    parse_card_format,
)
from tests.seed_data import MONTH, YEAR


# ---------------------------------------------------------------------------
# _clean_amount
# ---------------------------------------------------------------------------

class TestCleanAmount:
    @pytest.mark.parametrize("raw, expected", [
        ("-85,30",    -85.30),   # European comma decimal
        ("1.234,56",  1234.56),  # thousands dot + comma decimal
        ("500,00",    500.00),
        ("-0,99",     -0.99),
        ("100",       100.0),    # plain integer string
        ("100.50",    100.50),   # US decimal (no thousands dot)
    ])
    def test_valid_values(self, raw: str, expected: float) -> None:
        assert _clean_amount(raw) == pytest.approx(expected, abs=0.001)

    def test_none_on_nan(self) -> None:
        assert _clean_amount(float("nan")) is None

    def test_none_on_non_numeric_string(self) -> None:
        assert _clean_amount("not-a-number") is None

    def test_none_on_empty_string(self) -> None:
        assert _clean_amount("") is None


# ---------------------------------------------------------------------------
# _parse_date
# ---------------------------------------------------------------------------

class TestParseDate:
    @pytest.mark.parametrize("raw, expected", [
        ("03/03/2026",  date(2026, 3, 3)),
        ("2026-03-03",  date(2026, 3, 3)),
        ("03-03-2026",  date(2026, 3, 3)),
    ])
    def test_valid_formats(self, raw: str, expected: date) -> None:
        assert _parse_date(raw) == expected

    def test_passthrough_for_date_object(self) -> None:
        d = date(2026, 3, 15)
        assert _parse_date(d) == d

    def test_passthrough_for_timestamp(self) -> None:
        ts = pd.Timestamp("2026-03-15")
        assert _parse_date(ts) == date(2026, 3, 15)

    def test_none_on_nan(self) -> None:
        assert _parse_date(float("nan")) is None

    def test_none_on_unrecognised_format(self) -> None:
        assert _parse_date("15 March 2026") is None


# ---------------------------------------------------------------------------
# detect_format
# ---------------------------------------------------------------------------

class TestDetectFormat:
    def test_detects_account_format(self) -> None:
        df = pd.DataFrame(columns=[
            "Fecha de operación", "Fecha valor", "Concepto", "Importe", "Divisa", "Saldo"
        ])
        assert detect_format(df) == "account"

    def test_detects_card_format(self) -> None:
        df = pd.DataFrame(columns=[
            "Fecha operación", "Hora", "Nombre comercio", "Concepto", "Importe", "Divisa"
        ])
        assert detect_format(df) == "card"

    def test_raises_on_unknown_format(self) -> None:
        df = pd.DataFrame(columns=["Col1", "Col2", "Col3"])
        with pytest.raises(ValueError, match="Unrecognised"):
            detect_format(df)


# ---------------------------------------------------------------------------
# parse_account_format / parse_card_format
# ---------------------------------------------------------------------------

class TestParseAccountFormat:
    def _make_df(self) -> pd.DataFrame:
        return pd.DataFrame({
            "Fecha de operación": ["03/03/2026", "10/03/2026"],
            "Fecha valor":        ["03/03/2026", "10/03/2026"],
            "Concepto":           ["Supermercado", "Farmacia"],
            "Importe":            ["-62,50", "-18,90"],
            "Divisa":             ["EUR", "EUR"],
            "Saldo":              ["1000,00", "937,50"],
        })

    def test_output_columns(self) -> None:
        df = parse_account_format(self._make_df())
        assert set(df.columns) == {"date", "description", "amount", "source"}

    def test_source_is_account(self) -> None:
        df = parse_account_format(self._make_df())
        assert (df["source"] == "account").all()

    def test_amounts_parsed(self) -> None:
        df = parse_account_format(self._make_df())
        assert list(df["amount"]) == pytest.approx([-62.50, -18.90], abs=0.001)

    def test_drops_rows_with_null_amount(self) -> None:
        bad = self._make_df()
        bad.loc[0, "Importe"] = "n/a"
        df = parse_account_format(bad)
        assert len(df) == 1


class TestParseCardFormat:
    def _make_df(self) -> pd.DataFrame:
        return pd.DataFrame({
            "Fecha operación": ["15/03/2026"],
            "Hora":            ["12:30"],
            "Nombre comercio": ["Zara"],
            "Importe":         ["-45,00"],
            "Divisa":          ["EUR"],
        })

    def test_source_is_card(self) -> None:
        df = parse_card_format(self._make_df())
        assert (df["source"] == "card").all()

    def test_amount_parsed(self) -> None:
        df = parse_card_format(self._make_df())
        assert df["amount"].iloc[0] == pytest.approx(-45.00, abs=0.001)


# ---------------------------------------------------------------------------
# parse_bank_file (end-to-end via synthetic .xlsx fixture)
# ---------------------------------------------------------------------------

class TestParseBankFile:
    def test_returns_correct_row_count(self, account_xlsx_bytes: bytes) -> None:
        # 5 March rows + 1 February row; only 5 should survive the month filter
        df, fmt = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        assert len(df) == 5

    def test_returns_account_format(self, account_xlsx_bytes: bytes) -> None:
        _, fmt = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        assert fmt == "account"

    def test_all_rows_have_correct_month_year(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        assert (df["month"] == MONTH).all()
        assert (df["year"] == YEAR).all()

    def test_user_assigned(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Laerke", MONTH, YEAR)
        assert (df["user"] == "Laerke").all()

    def test_category_defaults_to_uncategorized(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        assert (df["category"] == "uncategorized").all()

    def test_date_is_iso_string(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        # ISO format: YYYY-MM-DD
        assert df["date"].iloc[0] == "2026-03-03"

    def test_dedup_drops_exact_duplicate_rows(self) -> None:
        """Two identical rows in the same file must result in one row after dedup."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(10):
            ws.append([""] * 5)
        ws.append(["Fecha de operación", "Fecha valor", "Concepto", "Importe",
                   "Divisa", "Saldo", "Nº mov", "Oficina"])
        # Same row twice
        for _ in range(2):
            ws.append(["05/03/2026", "05/03/2026", "Duplicate", "-10,00", "EUR",
                       "990,00", "1", "001"])
        buf = io.BytesIO()
        wb.save(buf)

        df, _ = parse_bank_file(buf.getvalue(), "dup.xlsx", "Hector", MONTH, YEAR)
        assert len(df) == 1


# ---------------------------------------------------------------------------
# df_to_records
# ---------------------------------------------------------------------------

class TestDfToRecords:
    def test_returns_list_of_dicts(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        records = df_to_records(df)
        assert isinstance(records, list)
        assert all(isinstance(r, dict) for r in records)

    def test_record_keys(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        records = df_to_records(df)
        expected_keys = {"user", "month", "year", "date", "description",
                         "amount", "source", "category"}
        assert set(records[0].keys()) == expected_keys

    def test_record_count_matches_df(self, account_xlsx_bytes: bytes) -> None:
        df, _ = parse_bank_file(account_xlsx_bytes, "test.xlsx", "Hector", MONTH, YEAR)
        records = df_to_records(df)
        assert len(records) == len(df)


# ---------------------------------------------------------------------------
# parse_bank_file_bulk (multi-month parsing)
# ---------------------------------------------------------------------------

class TestParseBankFileBulk:
    def test_returns_all_rows_no_month_filter(self, multi_month_xlsx_bytes: bytes) -> None:
        """Bulk parser should return rows from ALL months, not filter to a single month."""
        from budget.importer import parse_bank_file_bulk
        df, fmt = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        # 3 Jan + 3 Feb + 4 Mar = 10 total
        assert len(df) == 10

    def test_returns_bulk_format(self, multi_month_xlsx_bytes: bytes) -> None:
        from budget.importer import parse_bank_file_bulk
        _, fmt = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        assert fmt == "account"

    def test_month_year_derived_from_date(self, multi_month_xlsx_bytes: bytes) -> None:
        """Month and year should be extracted from each row's date."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")

        # Check that we have rows from 3 different months
        months = df[["month", "year"]].drop_duplicates()
        assert len(months) == 3

        # Verify each month
        for idx, row in months.iterrows():
            month, year = row["month"], row["year"]
            assert year == 2026
            assert month in [1, 2, 3]

    def test_january_transactions(self, multi_month_xlsx_bytes: bytes) -> None:
        """Verify January transactions are correctly parsed."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        jan_df = df[df["month"] == 1]
        assert len(jan_df) == 3
        assert (jan_df["year"] == 2026).all()

    def test_february_transactions(self, multi_month_xlsx_bytes: bytes) -> None:
        """Verify February transactions are correctly parsed."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        feb_df = df[df["month"] == 2]
        assert len(feb_df) == 3
        assert (feb_df["year"] == 2026).all()

    def test_march_transactions(self, multi_month_xlsx_bytes: bytes) -> None:
        """Verify March transactions are correctly parsed."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        mar_df = df[df["month"] == 3]
        assert len(mar_df) == 4
        assert (mar_df["year"] == 2026).all()

    def test_user_assigned(self, multi_month_xlsx_bytes: bytes) -> None:
        """User should be assigned to all rows."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Laerke")
        assert (df["user"] == "Laerke").all()

    def test_category_defaults_to_uncategorized(self, multi_month_xlsx_bytes: bytes) -> None:
        """All rows should default to uncategorized category."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        assert (df["category"] == "uncategorized").all()

    def test_date_is_iso_string(self, multi_month_xlsx_bytes: bytes) -> None:
        """Dates should be converted to ISO format strings (YYYY-MM-DD)."""
        from budget.importer import parse_bank_file_bulk
        df, _ = parse_bank_file_bulk(multi_month_xlsx_bytes, "multi.xlsx", "Hector")
        # Spot check: first January entry
        jan_df = df[df["month"] == 1]
        assert jan_df["date"].iloc[0] == "2026-01-05"

    def test_empty_file_returns_empty_df(self) -> None:
        """Empty file should return empty DataFrame."""
        from budget.importer import parse_bank_file_bulk
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(10):
            ws.append([""] * 5)
        ws.append(["Fecha de operación", "Fecha valor", "Concepto", "Importe",
                   "Divisa", "Saldo", "Nº mov", "Oficina"])
        buf = io.BytesIO()
        wb.save(buf)

        df, _ = parse_bank_file_bulk(buf.getvalue(), "empty.xlsx", "Hector")
        assert len(df) == 0

    def test_dedup_drops_duplicate_within_file(self) -> None:
        """Duplicates within the file should be dropped."""
        from budget.importer import parse_bank_file_bulk
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(10):
            ws.append([""] * 5)
        ws.append(["Fecha de operación", "Fecha valor", "Concepto", "Importe",
                   "Divisa", "Saldo", "Nº mov", "Oficina"])
        # Same row twice
        for _ in range(2):
            ws.append(["05/03/2026", "05/03/2026", "Duplicate", "-10,00", "EUR",
                       "990,00", "1", "001"])
        buf = io.BytesIO()
        wb.save(buf)

        df, _ = parse_bank_file_bulk(buf.getvalue(), "dup.xlsx", "Hector")
        assert len(df) == 1

    def test_cross_year_file(self) -> None:
        """File spanning multiple years should be parsed correctly."""
        from budget.importer import parse_bank_file_bulk
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(10):
            ws.append([""] * 5)
        ws.append(["Fecha de operación", "Fecha valor", "Concepto", "Importe",
                   "Divisa", "Saldo", "Nº mov", "Oficina"])
        # December 2025 + January 2026
        ws.append(["15/12/2025", "15/12/2025", "Dec transaction", "-50,00", "EUR", "1000,00", "1", "001"])
        ws.append(["05/01/2026", "05/01/2026", "Jan transaction", "-30,00", "EUR", "970,00", "2", "001"])
        buf = io.BytesIO()
        wb.save(buf)

        df, _ = parse_bank_file_bulk(buf.getvalue(), "cross_year.xlsx", "Hector")
        assert len(df) == 2

        # Check years
        years = sorted(df["year"].unique())
        assert years == [2025, 2026]

        # Check months
        dec_row = df[df["year"] == 2025]
        jan_row = df[df["year"] == 2026]
        assert dec_row["month"].iloc[0] == 12
        assert jan_row["month"].iloc[0] == 1

    def test_unrecognised_format_raises_error(self) -> None:
        """Unrecognised file format should raise ValueError."""
        from budget.importer import parse_bank_file_bulk
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(10):
            ws.append([""] * 5)
        ws.append(["Col1", "Col2", "Col3", "Col4"])  # Wrong columns
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="Unrecognised"):
            parse_bank_file_bulk(buf.getvalue(), "bad.xlsx", "Hector")
